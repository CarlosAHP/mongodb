import pymongo
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference

# Conectar a la base de datos MongoDB
client = pymongo.MongoClient("mongodb://localhost:27017/")

# Acceder a la base de datos 'SocioslegalesSA'
db = client["SocioslegalesSA"]

# Verificar la conexión
try:
    db.command("ping")
    print("Conexión exitosa a MongoDB")
except Exception as e:
    print("Error al conectar a MongoDB:", e)

# Función para extraer datos de una colección
def extraer_datos_coleccion(nombre_coleccion):
    collection = db.get_collection(nombre_coleccion)
    datos = list(collection.find({}))
    return datos

# Función para limpiar los datos
def limpiar_datos(datos):
    for registro in datos:
        for clave, valor in registro.items():
            if isinstance(valor, dict):
                if '$oid' in valor:
                    registro[clave] = valor['$oid']
                elif '$date' in valor:
                    registro[clave] = valor['$date']
    return datos

# Función para guardar datos en una hoja del archivo Excel
def guardar_datos_en_excel(nombre_coleccion, datos, writer):
    # Limpiar los datos
    datos_limpios = limpiar_datos(datos)
    
    # Crear un DataFrame de pandas
    df = pd.DataFrame(datos_limpios)
    
    # Eliminar el campo '_id' ya que no es necesario en el archivo Excel
    if '_id' in df.columns:
        df = df.drop(columns=['_id'])
    
    # Agregar una columna indicando la colección de origen
    df['Colección'] = nombre_coleccion
    
    # Reordenar las columnas para que 'Colección' esté al principio
    columnas = ['Colección'] + [col for col in df.columns if col != 'Colección']
    df = df[columnas]
    
    # Guardar el DataFrame en una hoja del archivo Excel
    df.to_excel(writer, sheet_name=nombre_coleccion, index=False)
    
    return df

# Crear un escritor de pandas para manejar múltiples hojas
archivo_excel = "SocioslegalesSA_datos_nuevo_4.xlsx"  # Cambiar el nombre del archivo
writer = pd.ExcelWriter(archivo_excel, engine='openpyxl')

# Lista para almacenar los DataFrames de cada colección
dataframes = []

# Extraer y guardar datos de la colección 'Cliente'
datos_cliente = extraer_datos_coleccion("Cliente")
df_cliente = guardar_datos_en_excel("Cliente", datos_cliente, writer)
dataframes.append(df_cliente)

# Extraer y guardar datos de la colección 'Abogado'
datos_abogado = extraer_datos_coleccion("Abogado")
df_abogado = guardar_datos_en_excel("Abogado", datos_abogado, writer)
dataframes.append(df_abogado)

# Extraer y guardar datos de la colección 'Asunto'
datos_asunto = extraer_datos_coleccion("Asunto")
df_asunto = guardar_datos_en_excel("Asunto", datos_asunto, writer)
dataframes.append(df_asunto)

# Extraer y guardar datos de la colección 'Procurador'
datos_procurador = extraer_datos_coleccion("Procurador")
df_procurador = guardar_datos_en_excel("Procurador", datos_procurador, writer)
dataframes.append(df_procurador)

# Extraer y guardar datos de la colección 'Audiencia'
datos_audiencia = extraer_datos_coleccion("Audiencia")
df_audiencia = guardar_datos_en_excel("Audiencia", datos_audiencia, writer)
dataframes.append(df_audiencia)

# Crear una hoja maestra con todos los datos combinados
df_maestra = pd.concat(dataframes, ignore_index=True)
df_maestra.to_excel(writer, sheet_name="Maestra", index=False)

# Guardar y cerrar el archivo Excel
writer.close()

# Abrir el archivo Excel para añadir gráficos
wb = load_workbook(archivo_excel)
ws = wb["Maestra"]

# Contar el número de asuntos por estado
# Buscar la columna 'estado' en la hoja maestra
estado_col = None
for col in ws.iter_cols(1, ws.max_column):
    if col[0].value == 'estado':
        estado_col = col[0].column_letter
        break

if estado_col:
    estado_counts = {}
    for cell in ws[estado_col][1:]:
        if cell.value:
            if cell.value in estado_counts:
                estado_counts[cell.value] += 1
            else:
                estado_counts[cell.value] = 1

    # Agregar los datos de conteo de estados a la hoja maestra
    start_row = ws.max_row + 2  # Añadir después de los datos existentes
    ws.cell(row=start_row, column=10, value="Estado")
    ws.cell(row=start_row, column=11, value="Cantidad")
    for i, (estado, count) in enumerate(estado_counts.items(), start=start_row + 1):
        ws.cell(row=i, column=10, value=estado)
        ws.cell(row=i, column=11, value=count)

    # Crear un gráfico de barras
    bar_chart = BarChart()
    data = Reference(ws, min_col=11, min_row=start_row, max_col=11, max_row=start_row + len(estado_counts))
    categories = Reference(ws, min_col=10, min_row=start_row + 1, max_row=start_row + len(estado_counts))
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(categories)
    bar_chart.title = "Número de Asuntos por Estado"
    bar_chart.y_axis.title = 'Cantidad'
    bar_chart.x_axis.title = 'Estado'

    # Ubicar el gráfico en la hoja de cálculo
    ws.add_chart(bar_chart, "J2")

# Contar el número de asuntos por gabinete
# Buscar la columna 'gabinete' en la hoja maestra
gabinete_col = None
for col in ws.iter_cols(1, ws.max_column):
    if col[0].value == 'gabinete':
        gabinete_col = col[0].column_letter
        break

if gabinete_col:
    gabinete_counts = {}
    for cell in ws[gabinete_col][1:]:
        if cell.value:
            if cell.value in gabinete_counts:
                gabinete_counts[cell.value] += 1
            else:
                gabinete_counts[cell.value] = 1

    # Agregar los datos de conteo de gabinetes a la hoja maestra
    start_row = ws.max_row + 2  # Añadir después de los datos existentes
    ws.cell(row=start_row, column=10, value="Gabinete")
    ws.cell(row=start_row, column=11, value="Cantidad")
    for i, (gabinete, count) in enumerate(gabinete_counts.items(), start=start_row + 1):
        ws.cell(row=i, column=10, value=gabinete)
        ws.cell(row=i, column=11, value=count)

    # Crear un gráfico de pastel
    pie_chart_gabinete = PieChart()
    labels = Reference(ws, min_col=10, min_row=start_row + 1, max_row=start_row + len(gabinete_counts))
    data = Reference(ws, min_col=11, min_row=start_row, max_col=11, max_row=start_row + len(gabinete_counts))
    pie_chart_gabinete.add_data(data, titles_from_data=True)
    pie_chart_gabinete.set_categories(labels)
    pie_chart_gabinete.title = "Distribución por Gabinete"

    # Ubicar el gráfico en la hoja de cálculo
    ws.add_chart(pie_chart_gabinete, "R2")

# Contar los nombres más comunes en los casos
nombre_col = None
for col in ws.iter_cols(1, ws.max_column):
    if col[0].value == 'nombre':
        nombre_col = col[0].column_letter
        break

if nombre_col:
    nombre_counts = {}
    for cell in ws[nombre_col][1:]:
        if cell.value:
            if cell.value in nombre_counts:
                nombre_counts[cell.value] += 1
            else:
                nombre_counts[cell.value] = 1

    # Agregar los datos de conteo de nombres a la hoja maestra
    start_row = ws.max_row + 2  # Añadir después de los datos existentes
    ws.cell(row=start_row, column=10, value="Nombre")
    ws.cell(row=start_row, column=11, value="Cantidad")
    for i, (nombre, count) in enumerate(nombre_counts.items(), start=start_row + 1):
        ws.cell(row=i, column=10, value=nombre)
        ws.cell(row=i, column=11, value=count)

    # Crear un gráfico de pastel para los nombres más comunes
    pie_chart_nombre = PieChart()
    labels = Reference(ws, min_col=10, min_row=start_row + 1, max_row=start_row + len(nombre_counts))
    data = Reference(ws, min_col=11, min_row=start_row, max_col=11, max_row=start_row + len(nombre_counts))
    pie_chart_nombre.add_data(data, titles_from_data=True)
    pie_chart_nombre.set_categories(labels)
    pie_chart_nombre.title = "Nombres Más Comunes en Casos"

    # Ubicar el gráfico en la hoja de cálculo
    ws.add_chart(pie_chart_nombre, "R20")

# Guardar los cambios en el archivo Excel con gráficos
wb.save(archivo_excel)
print(f"Datos guardados en '{archivo_excel}' con gráficos.")
